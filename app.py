from flask import Flask, jsonify,Response, request, make_response, send_from_directory
from flask_cors import CORS
import os, uuid, io, socket
from datetime import timedelta, datetime
from flask_bcrypt import Bcrypt
from flask_migrate import Migrate
from werkzeug.utils import secure_filename
from flask_jwt_extended import JWTManager, create_access_token, jwt_required, get_jwt_identity
from models import db, Users, Profiles, Users_profiles, Type, Model, Brand, State, Pc, Employes, Employes_state, Log, Brandcell, Cell
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
####prueba de datos
from faker import Faker
fake = Faker()

#####

app = Flask(__name__)

# Configuración base de datos myql
#app.config['SQLALCHEMY_DATABASE_URI'] = 'mysql://root@localhost/afn'

app.config['SQLALCHEMY_DATABASE_URI'] = 'mssql+pyodbc://afn:123@(localdb)\dbprueba/sistema_afn?driver=ODBC+Driver+17+for+SQL+Server'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
#folder pdf
app.config['UPLOAD_FOLDER'] = './pdf_afn'

# Configuración de JWT
app.config['JWT_SECRET_KEY'] = '9123'
app.config["JWT_ACCESS_TOKEN_EXPIRES"] = timedelta(hours=4)
jwt = JWTManager(app)

bcrypt = Bcrypt(app)
#supports_credentials permite el envio de cookies
CORS(app, supports_credentials=True)
migrate = Migrate(app, db)
db.init_app(app)

@app.route('/login', methods=['POST'])
def login():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        user = Users.query.filter_by(nick_name=username).first()

        if user and bcrypt.check_password_hash(user.u_password, password):
            # Crear un token de acceso
            access_token = create_access_token(identity=user.cod_user)
            response = make_response(jsonify(access_token=access_token), 200)
            response.set_cookie('access_token_cookie', value=access_token, httponly=True, samesite='None',secure=True)

            return response

    return jsonify({'error': 'Credenciales inválidas'}), 401

@app.route('/logout', methods=['POST'])
@jwt_required()
def logout():
    
    return jsonify({'message': 'Desconectado'}), 200

@app.route('/register', methods=['POST'])
@jwt_required()
def register():
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        profile = request.form['profile']
        existe_user = Users.query.filter_by(nick_name=username).first()

        if username == 'undefined' or password == 'undefined' or profile == "":
            return jsonify({'message': 'Todos los campos deben ingresados'}), 401     
           
        if existe_user:
            return jsonify({'message': 'El nombre de usuario ya está en uso. Elige otro.'}), 400
        

        if username != 'undefined':
            hashed_password = bcrypt.generate_password_hash(password).decode('utf-8')
            new_user = Users(nick_name=username, u_password=hashed_password)
            db.session.add(new_user)
            db.session.commit()

            add_profile = Users_profiles(cod_user_id=new_user.cod_user,cod_profile_id=profile)
            db.session.add(add_profile)
            db.session.commit()
            db.session.close() 
            return jsonify({'message': 'Usuario registrado exitosamente'}), 201

    return jsonify({'message': 'Registro fallido'}), 400

@app.route('/check-token', methods=['GET'])
@jwt_required()
def check_auth():
    current_user_id = get_jwt_identity()
    current_user = Users.query.filter_by(cod_user=current_user_id).first()
    #db.session.get
    user_profile = Users_profiles.query.filter_by(cod_user_id=current_user.cod_user).first()

    user_profile_details = {
        'cod_user' : user_profile.cod_user_id,
        'name_user': current_user.nick_name,
        'profile' : user_profile.cod_profile_id
    }
    return jsonify({'authenticated': True, 'user_profile': user_profile_details})

@app.route('/add_employes', methods= ['POST'])
@jwt_required()
def add_user_pc():

    if request.method == 'POST':
        lastname_user = request.form['lastname_user']
        create_date = request.form['create_date']
        gps = request.form['gps']

        ##cambiar fecha a sqlserver...
        new_create_date = datetime.strptime(create_date,'%Y-%m-%dT%H:%M')

        sql_fecha = new_create_date.strftime('%Y-%m-%d %H:%M:%S')
        
        existe_gps = Employes.query.filter_by(gps_id=gps).first()

        if  existe_gps:
            return jsonify({"estado":"El GPS id ya esta en uso."}), 400
        
        if gps and  lastname_user == 'undefined':
            return jsonify({'mensaje':'Debe ingresar los los campos de GPS y Nombre completo'}),402

        if gps:
            #agregar empleado
            new_user_pc = Employes(lastname_user=lastname_user, create_date=sql_fecha, gps_id=gps,cod_employe_id=1)
            db.session.add(new_user_pc)
            db.session.commit()
            
            #registrar Log
            db.session.close() 
            return jsonify({"mensaje":"usuario creado"}), 200
        else:
            return jsonify({"mensaje":"gps faltante"}, 401)

       
    return jsonify({"estado":"Error al agregar"})

@app.route('/edit_employes' , methods=['POST'])
@jwt_required()
def edit_employe():

    if request.method == 'POST':

        cod = request.form['cod_pusers']
        gps = request.form['gps_id']
        lastname = request.form['lastname']
        service_tag = request.form['service_tag']
        archivo = request.files.get('pdf_file')
        date_delivery = request.form['date']


        current_user_id = get_jwt_identity()
        current_user = Users.query.get(current_user_id)

        data_pc = Pc.query.filter_by(service_tag=service_tag).first()
        data_employes = Employes.query.filter_by(gps_id=gps).first()
        cod_usuario = Employes.query.filter_by(cod_employes=cod).first()

        #cambiar fecha a sqlserver...
        new_date = datetime.strptime(date_delivery,'%Y-%m-%dT%H:%M')

        sql_fecha = new_date.strftime('%Y-%m-%d %H:%M:%S')


        if service_tag and service_tag != '0':
            
            if archivo:
                # tomar archivo y poner otro nombre
                filename = secure_filename(archivo.filename)
                unique_filename = str(uuid.uuid4())+'_'+filename
                archivo.save(os.path.join(app.config['UPLOAD_FOLDER'],unique_filename))
                    
                #editar estado de pc. 1 = asignado
                data_pc.cod_state_id = 1
                db.session.commit()
                #asignar pc a empleado
                data_employes.cod_pc_id = data_pc.cod_pc    
                db.session.commit()
                #añadir nombre de archivo 
                data_employes.archivo = unique_filename
                db.session.commit()
                #añadir fecha de entrega
                data_employes.date_delivery = sql_fecha
                db.session.commit()
            else:
                
                #editar estado de pc. 1 = asignado
                data_pc.cod_state_id = 1
                db.session.commit()
                #asignar pc a empleado
                data_employes.cod_pc_id = data_pc.cod_pc    
                db.session.commit()
                #añadir fecha de entrega
                data_employes.date_delivery = sql_fecha
                db.session.commit()

                data_employes.archivo = None
                db.session.commit()

            if cod_usuario:
                if cod_usuario.lastname_user != lastname or cod_usuario.gps_id != gps:

                    cod_usuario.lastname_user = lastname
                    db.session.commit()

                    if data_employes is None:
                        cod_usuario.gps_id = gps
                        db.session.commit()

            #Registrar LOG - asignar pc
            state_asignado = State.query.filter_by(name_state='asignado').first_or_404()
            new_log_asigar =  Log(log_pc_id=data_pc.cod_pc,log_pc_nc=data_pc.name_computer,log_pc_st=data_pc.service_tag,log_date=sql_fecha,
                                log_archivo=data_employes.archivo,log_cod_user_id=current_user.cod_user,log_cod_employe=data_employes.gps_id,
                                log_name_employe=data_employes.lastname_user,log_state=state_asignado.name_state
                                )
            db.session.add(new_log_asigar)
            db.session.commit()
            db.session.close()
            return jsonify({"mensaje":"Asigando Exitosamente"}),205
        
        if cod_usuario:
            if cod_usuario.lastname_user != lastname or cod_usuario.gps_id != gps:
                if data_employes is None:
                    cod_usuario.gps_id = gps
                    db.session.commit()
                    cod_usuario.lastname_user = lastname
                    db.session.commit()
                    return jsonify({"mensaje":"datos actualizados"}),200
                elif cod_usuario.lastname_user != lastname:
                    cod_usuario.lastname_user = lastname
                    db.session.commit()
                    return jsonify({"mensaje":"datos actualizados"}),200
                elif data_employes:
                    return jsonify({"mensaje":"gps existe"}),403
            else:
                return jsonify({"estado":"no actualizar"}),401
            
    return jsonify({"estado":"error de conexion"}),400

@app.route('/reasignar_pc', methods = ['POST'])
@jwt_required()
def reasigar():
    if request.method == 'POST':

        cod = request.form['cod_pusers']
        gps = request.form['gps_id']
        lastname = request.form['lastname']
        st = request.form['service_tag']
        state = request.form['state']
        new_st = request.form['new_st']
        archivo = request.files.get('pdf_file')
        date_delivery = request.form['date']
        view = request.form['view']

        existe_st = Pc.query.filter_by(service_tag=st).first()
        cod_usuario = Employes.query.filter_by(cod_employes=cod).first()
        existe_new_st = Pc.query.filter_by(service_tag=new_st).first()
        data_employes = Employes.query.filter_by(gps_id=gps).first()
        buscar_id_employe = Employes.query.filter_by(cod_pc_id = existe_st.cod_pc).first()

        current_user_id = get_jwt_identity()
        current_user = Users.query.get(current_user_id)

        ##cambiar fecha a sqlserver...
        new_create_date = datetime.strptime(date_delivery,'%Y-%m-%dT%H:%M')

        sql_fecha = new_create_date.strftime('%Y-%m-%d %H:%M:%S')

        #Reasignar
        if view == '1':
            if existe_st and existe_new_st:
                if buscar_id_employe:
                    # Cambiar estado a pc actual
                    existe_st.cod_state_id = state
                    db.session.commit()
                    #LOG-CAMBIAR ESTADO PC ACTUAL
                    buscar_name_state = State.query.filter_by(cod_state=state).first_or_404()
                    new_log_cambiar_estado = Log(
                        log_pc_id=existe_st.cod_pc,log_pc_nc=existe_st.name_computer,log_pc_st=existe_st.service_tag,
                        log_date=sql_fecha,log_state=buscar_name_state.name_state,log_cod_user_id=current_user.cod_user,
                        log_cod_employe=buscar_id_employe.gps_id,log_name_employe=buscar_id_employe.lastname_user,
                        log_archivo=buscar_id_employe.archivo
                        )
                    db.session.add(new_log_cambiar_estado)
                    db.session.commit()
                    # Cambiar estado a pc nuevo
                    existe_new_st.cod_state_id = 1
                    db.session.commit()
                    # Asignar pc nuevo
                    buscar_id_employe.cod_pc_id = existe_new_st.cod_pc
                    db.session.commit()
                    # Cambiar fecha
                    buscar_id_employe.date_delivery = sql_fecha
                    db.session.commit()


                    if archivo:
                        #asignar archivo
                        filename = secure_filename(archivo.filename)
                        unique_filename = str(uuid.uuid4())+'_'+filename
                        archivo.save(os.path.join(app.config['UPLOAD_FOLDER'],unique_filename))

                        buscar_id_employe.archivo = unique_filename
                        db.session.commit()
                    else:
                        # archivo vacio
                        buscar_id_employe.archivo = None
                        db.session.commit()

                    if cod_usuario.lastname_user != lastname or cod_usuario.gps_id != gps:
                        cod_usuario.lastname_user = lastname
                        db.session.commit()
                        jsonify({"mensaje":"nombre actualizados"})

                        if data_employes is None:
                            cod_usuario.gps_id = gps
                            db.session.commit()
                            jsonify({"mensaje":"gps actualizados"})

                    #LOG-REASIGNAR
                    state_asignado = State.query.filter_by(name_state='asignado').first_or_404()
                    new_log_reasignar = Log(log_pc_id=existe_new_st.cod_pc,log_pc_nc=existe_new_st.name_computer,log_pc_st=existe_new_st.service_tag,
                                            log_date=sql_fecha,log_archivo=buscar_id_employe.archivo,log_cod_user_id=current_user.cod_user,
                                            log_cod_employe=buscar_id_employe.gps_id,log_name_employe=buscar_id_employe.lastname_user,log_state=state_asignado.name_state                                       
                                           )
                    db.session.add(new_log_reasignar)
                    db.session.commit()
                    db.session.close()



                    return jsonify({"mensaje":"computador reasignado"}),200
            return jsonify({"mensaje":"no hay computadores para asignaar"}),402
        elif view == '2':
            
            if cod_usuario.lastname_user != lastname or cod_usuario.gps_id != gps:
                cod_usuario.lastname_user = lastname
                db.session.commit()
                jsonify({"mensaje":"nombre actualizados"})
                if data_employes is None:
                    cod_usuario.gps_id = gps
                    db.session.commit()
                    jsonify({"mensaje":"gps actualizados"})

            #Log-Quitar
            state_asignado = State.query.filter_by(cod_state=state).first_or_404()
            new_log_quitar = Log(log_pc_id=existe_st.cod_pc,log_pc_nc=existe_st.name_computer,log_pc_st=existe_st.service_tag,
                                log_date=sql_fecha,log_state='removido',log_archivo=buscar_id_employe.archivo,
                                log_cod_user_id=current_user.cod_user,log_cod_employe=buscar_id_employe.gps_id,
                                log_name_employe=buscar_id_employe.lastname_user
                                 )
            db.session.add(new_log_quitar)
            db.session.commit()
            # Cambiar estado a pc
            existe_st.cod_state_id = state 
            db.session.commit()

            # quitar pc a empleado
            buscar_id_employe.cod_pc_id = None
            db.session.commit()

            # limpiar pdf 
            buscar_id_employe.archivo = None
            db.session.commit()

            #limpiar fecha
            buscar_id_employe.date_delivery = None
            db.session.commit()


            db.session.close()
            return jsonify({"mensaje":"computador quitado"}),201      
        elif view == '3':
            #cambiar o editar archivo

            if cod_usuario.lastname_user != lastname or cod_usuario.gps_id != gps:
                cod_usuario.lastname_user = lastname
                db.session.commit()
                jsonify({"mensaje":"nombre actualizados"})
                if data_employes is None:
                    cod_usuario.gps_id = gps
                    db.session.commit()
                    jsonify({"mensaje":"gps actualizados"})

            if archivo and buscar_id_employe:
                #asignar archivo
                filename = secure_filename(archivo.filename)
                unique_filename = str(uuid.uuid4())+'_'+filename
                archivo.save(os.path.join(app.config['UPLOAD_FOLDER'],unique_filename))

                buscar_id_employe.archivo = unique_filename
                db.session.commit()
                db.session.close()
                return jsonify({"mensaje":"archivo actualizado"})
            else:
                # archivo vacio
                buscar_id_employe.archivo = None
                db.session.commit()
                db.session.close()
                return jsonify({"mensaje":"computador reasignado"}),203
        elif cod_usuario.lastname_user != lastname or cod_usuario.gps_id != gps:
            cod_usuario.lastname_user = lastname
            db.session.commit()

            if data_employes is None:
                cod_usuario.gps_id = gps
                db.session.commit()
            jsonify({"mensaje":"nombre o gps actualizados"})
        else:
            return jsonify({"mensaje":"debe seleccionar una opción"}),401
    return jsonify({"mensaje":"error al actualizar"})

@app.route('/uploads/<name>')
def download_file(name):
    return send_from_directory(app.config["UPLOAD_FOLDER"], name)

@app.route('/add_pc', methods = ['POST'])
@jwt_required()
def create():
    if request.method == 'POST':
        name_computer = request.form['namecomputer']
        st = request.form['nameserial']
        date_received = request.form['datereceived']
        cod_model_id = request.form['namemodel']
        cod_brand_id = request.form['namebrand']
        cod_type_id = request.form['nametype']

        ##cambiar fecha a sqlserver...
        new_date = datetime.strptime(date_received,'%Y-%m-%dT%H:%M')

        sql_fecha = new_date.strftime('%Y-%m-%d %H:%M:%S')
    
        if name_computer == 'undefined':
            return jsonify({"mensaje":"campo nombre vacio"}),403
        if st == 'undefined':
            return jsonify({"mensaje":"campo st vacio"}),404

        if name_computer and st and date_received and cod_brand_id and cod_model_id and cod_type_id:
            current_user_id = get_jwt_identity()
            current_user = Users.query.get(current_user_id)

            existe_service_tag = Pc.query.filter_by(service_tag=st).first()
            existe_name_computer = Pc.query.filter_by(name_computer=name_computer).first()
            if existe_service_tag and existe_name_computer:
                return jsonify({'message': 'El nombre de numero de serial y nombre computador ya está en uso.'}), 400
            elif existe_name_computer:
                return jsonify({'message': 'El nombre de computador ya está en uso.'}), 401
            elif existe_service_tag:
                return  jsonify({'message': 'El numero de serial ya está en uso.'}), 402
            

            new_pc = Pc(name_computer=name_computer,service_tag=st,date_received=sql_fecha,cod_model_id=cod_model_id,cod_brand_id=cod_brand_id,cod_type_id=cod_type_id,cod_user_id=current_user.cod_user,cod_state_id=2)
            db.session.add(new_pc)
            db.session.commit()
            
            #LOG-ADD
            s_name = State.query.filter_by(name_state='disponible').first_or_404()
            new_log = Log(log_pc_id=new_pc.cod_pc,log_pc_nc=name_computer,log_pc_st=st,log_state=s_name.name_state,log_date=sql_fecha,log_cod_user_id=current_user.cod_user)
            db.session.add(new_log)
            db.session.commit()

            db.session.close() 
            return jsonify({'pc':'Creado'})
        


    return jsonify({'pc':'Error ingreso'})

@app.route('/edit_pc', methods = ['POST'])
@jwt_required()
def edit_pc():

    if request.method == 'POST':
        cod = request.form['cod']
        computer = request.form['computer']
        st = request.form['st']
        marca = request.form['marca']
        tipo = request.form['tipo']
        modelo = request.form['model']
        estado = request.form['estado']

        existe_cod = Pc.query.filter_by(cod_pc=cod).first()

        existe_st = Pc.query.filter_by(service_tag=st).first()

        existe_namepc = Pc.query.filter_by(name_computer=computer).first()

        if existe_cod:
            
            if existe_cod.service_tag == st:
                existe_cod.service_tag = st
                db.session.commit()

            elif existe_cod.service_tag != st and existe_st:
                return jsonify({"mensaje":"ST ya esta en uso"}),401
            else:
                existe_cod.service_tag = st
                db.session.commit()

            if existe_cod.name_computer == computer:
                existe_cod.name_computer = computer
                db.session.commit()

            elif existe_cod.name_computer != computer and existe_namepc:
                return jsonify({"mensaje":"NC ya esta en uso"}),402
            else:
                existe_cod.name_computer = computer
                db.session.commit()

            
            existe_cod.cod_brand_id = marca
            db.session.commit()
   
            existe_cod.cod_type_id = tipo
            db.session.commit()

            existe_cod.cod_model_id = modelo
            db.session.commit()

            existe_cod.cod_state_id = estado
            db.session.commit()

            #LOG-EDITAR 




            db.session.close()   

            return jsonify({"mensaje":"actualizado"})

    return jsonify({"mensaje":"error eliminar"}),403

@app.route('/delete_pc/<int:cod>', methods = ['DELETE'])
def delete_pc(cod):

    if request.method == 'DELETE':
        borrar = Pc.query.get_or_404(cod)

        db.session.delete(borrar)
        db.session.commit()
        db.session.close()
        
        return jsonify({"mensaje":"eliminado"})
    return jsonify({"mensaje":"error al eliminar"})

@app.route('/delete_employe/<int:cod>', methods = ['DELETE'])
def delete_employe(cod):
    if request.method == 'DELETE':
        borrar = Employes.query.get_or_404(cod)

        db.session.delete(borrar)
        db.session.commit()
        db.session.close()
        
        return jsonify({"mensaje":"eliminado"})
    return jsonify({"mensaje":"error al eliminar"})

@app.route('/add_brand', methods = ['POST'])
def add_brand():

    if request.method == 'POST':
        marca = request.form['marca']
        if marca != 'undefined':
            new_marca = Brand(name_brand = marca)
            db.session.add(new_marca)
            db.session.commit()
            db.session.close() 
            return jsonify({"mensaje":"marca añadida"})
        return jsonify({"mensaje":"debe ingresar nombre de marca"}),401
    return jsonify({"mensaje":"error al añadir marca"})

@app.route('/delete_brand', methods = ['POST'])
def delete_brand():
    if request.method== 'POST':
        cod = request.form['cod_brand']
        existe_brand = Brand.query.filter_by(cod_brand=cod).first()
        existe_pc_brand = Pc.query.filter_by(cod_brand_id=cod).first()
        if int(cod) != 0:
            if existe_pc_brand is None:   
                db.session.delete(existe_brand)
                db.session.commit()
                db.session.close() 
                return jsonify({"mensaje":"marca eliminada"})
            else:
                return jsonify({"mensaje":"no puede ser eliminada"}),402
        return jsonify({"mensaje":"Debe elegir marca"}), 401
    return  jsonify({"mensaje":"error al eliminar"})

@app.route('/add_model', methods = ['POST'])
def add_modelo():

    if request.method == 'POST':
        modelo = request.form['modelo']
        if modelo != 'undefined':
            new_modelo = Model(name_model = modelo)
            db.session.add(new_modelo)
            db.session.commit()
            db.session.close() 
            return jsonify({"mensaje":"modelo añadida"})
        return jsonify({"mensaje":"debe ingresar nombre de modelo"}),401
    return jsonify({"mensaje":"error al añadir modelo"})

@app.route('/delete_model', methods = ['POST'])
def delete_modelo():
    if request.method == 'POST':
        cod = request.form['cod_model']
        existe_modelo = Model.query.filter_by(cod_model=cod).first()
        existe_pc_modelo = Pc.query.filter_by(cod_model_id=cod).first()
        if int(cod) != 0:
            if existe_pc_modelo is None:   
                db.session.delete(existe_modelo)
                db.session.commit()
                db.session.close() 
                return jsonify({"mensaje":"modelo eliminada"})
            else:
                return jsonify({"mensaje":"no puede ser eliminada"}),402
        return jsonify({"mensaje":"Debe seleccionar una modelo para eliminar"}), 401
    return  jsonify({"mensaje":"error al eliminar"})

@app.route('/add_type', methods = ['POST'])
def add_tipo():

    if request.method == 'POST':
        tipo = request.form['tipo']
        if tipo != 'undefined':
            new_type = Type(name_type = tipo)
            db.session.add(new_type)
            db.session.commit()
            db.session.close() 
            return jsonify({"mensaje":"tipo añadida"})
        return jsonify({"mensaje":"debe ingresar nombre de tipo"}),401
    return jsonify({"mensaje":"error al añadir tipo"})

@app.route('/delete_type', methods = ['POST'])
def delete_tipo():
    if request.method == 'POST':
        cod = request.form['cod_type']
        existe_tipo = Type.query.filter_by(cod_type=cod).first()
        existe_pc_tipo = Pc.query.filter_by(cod_type_id=cod).first()
        if int(cod) != 0:
            if existe_pc_tipo is None:   
                db.session.delete(existe_tipo)
                db.session.commit()
                db.session.close() 
                return jsonify({"mensaje":"tipo eliminada"})
            else:
                return jsonify({"mensaje":"no puede ser eliminada"}),402
        return jsonify({"mensaje":"Debe seleccionar un tipo para eliminar"}), 401
    return  jsonify({"mensaje":"error al eliminar"})

@app.route('/add_cell', methods = ['POST'])
def add_cell():
    if request.method == 'POST':   
        number_cell = request.form['numero']
        imei = request.form['imei']
        cod = request.form['marca']
        exist_imei = Cell.query.filter_by(imei=imei).first()
        if number_cell and imei and cod:
            if exist_imei:
                return jsonify({"mensaje":"imei existe"}), 402
            else:
                buscar = State.query.filter_by(name_state='disponible').first()
                new_cell = Cell(imei=imei,number_cell=number_cell,cod_brand_cell_id=cod,cod_state_id=buscar.cod_state)
                db.session.add(new_cell)
                db.session.commit()
                db.session.close()
                return jsonify({"mensaje":"celular creado"})
        return jsonify({"mensaje":"error al agregar"}), 401
    return jsonify({"mensaje":"error conexion"}),400

@app.route('/get_cell', methods = ['GET'])
def get_cell():
    resultado = db.session.query(Cell.cod_cell,Cell.imei,Cell.number_cell,Cell.cod_state_id,Brandcell.name_brand_cell,Brandcell.cod_brand_cell,State.name_state)\
                .join(Brandcell, Cell.cod_brand_cell_id == Brandcell.cod_brand_cell)\
                .join(State, Cell.cod_state_id == State.cod_state)\
                .all()
   
    data = []

    for cell in resultado:
        data.append({
            'cod_cell': cell.cod_cell,
            'imei':cell.imei,
            'number_cell': cell.number_cell,
            'name_brand_cell': cell.name_brand_cell,
            'cod_brand_cell': cell.cod_brand_cell,
            'name_state': cell.name_state,
            'cod_state_id' : cell.cod_state_id
        })
    return jsonify(data)

@app.route('/delete_cell/<int:cod>', methods = ['DELETE'])
def delete_cell(cod):

    if request.method == 'DELETE':
        
        borrar = Cell.query.get_or_404(cod)

        db.session.delete(borrar)
        db.session.commit()
        db.session.close()
        
        return jsonify({"mensaje":"eliminado"})

    return jsonify({"mensaje":"error al eliminar"})

@app.route('/add_brandcell', methods = ['POST'])
def add_brandcell():
    if request.method == 'POST':
        brandcell = request.form['marca']
        if brandcell != 'undefined' and brandcell != '':
            new_brandcell = Brandcell(name_brand_cell=brandcell)
            db.session.add(new_brandcell)
            db.session.commit()
            db.session.close()
            return jsonify({"mensaje":"marca añadida"})
        return jsonify({"mensaje":"debe ingresar nombre de marca"}),401
    return jsonify({"mensaje":"error al añadir marca"})

@app.route('/delete_brandcell', methods = ['POST'])
def delete_brandcell():

    if request.method == 'POST':
        cod = request.form['marcae']
        existe_cod = Brandcell.query.filter_by(cod_brand_cell=cod).first()
        existe_codCell = Cell.query.filter_by(cod_brand_cell_id=cod).first()
        if cod != '':
            if existe_codCell is None:
                db.session.delete(existe_cod)
                db.session.commit()
                db.session.close()
                return jsonify({"mensaje":"marca eliminada"})
            return jsonify({"mensaje":"no puede ser eliminada"}),403
        return jsonify({"mensaje":"seleccionar opción"}),401
    return  jsonify({"mensaje":"error al eliminar"}),402

@app.route('/get_brandcell', methods = ['GET'])
def get_brandcell():

    brandcell =  Brandcell.query.all()
    models_brandcell = [brand.obtener() for brand in brandcell]

    return  jsonify(models_brandcell)

@app.route('/datos', methods = ['GET'])
def datos():

    models = Model.query.all()
    types = Type.query.all()
    brands = Brand.query.all()
    profiles =  Profiles.query.all()
    states = State.query.all()
    
    models_json = [model.obtener() for model in models]
    types_json = [tipo.obtener() for tipo in types]
    brand_json = [brand.obtener() for brand in brands]
    profiles_json = [profile.obtener() for profile in profiles]
    states_json = [state.obtener() for state in states]
    datos = {
        'model':models_json,
        'type':types_json,
        'brand': brand_json,
        'profile': profiles_json,
        'state': states_json
    }
    return jsonify(datos)    

@app.route('/estado', methods = ['GET'])
def estado():

    states = State.query.all()

    states_json = [state.obtener() for state in states]

    datos = {
        'state': states_json 
    }

    return jsonify(datos)

@app.route('/get_pc_users', methods = ['GET'])
def get_pc_users():

    resultado = db.session.query(Employes,Employes.cod_employes,Pc.service_tag,Pc.name_computer,Employes.archivo,Employes.date_delivery, Employes.gps_id,Employes.lastname_user,Employes.create_date,Employes_state.state_employe)\
                .join(Employes_state, Employes.cod_employe_id == Employes_state.cod_employe_state)\
                .join(Pc, Employes.cod_pc_id == Pc.cod_pc, isouter=True)\
                .order_by(
                    db.case(
                        (Employes.cod_pc_id == None,0),
                        else_=1
                        ))\
                .all()

                

    data = []
    for user in resultado:
        data.append(
            {
                "cod_pusers" : user.cod_employes,
                "lastname_user": user.lastname_user,
                "create_date" : user.create_date,
                "gps_id" : user.gps_id,
                "service_tag" : user.service_tag,
                "name_computer" : user.name_computer,
                "archivo" : user.archivo,
                "date_delivery" : user.date_delivery
            }
        )

    return jsonify(data)

@app.route('/get_pc', methods = ['GET','POST'])
def get_pc():

    resultado = db.session.query(Pc,Pc.cod_pc,Pc.name_computer,Pc.service_tag,Pc.date_received,Employes.lastname_user,Brand.cod_brand,Brand.name_brand,Model.cod_model,Model.name_model,Type.cod_type,Type.name_type,State.cod_state,State.name_state)\
                .join(Brand, Pc.cod_brand_id == Brand.cod_brand)\
                .join(Model, Pc.cod_model_id == Model.cod_model)\
                .join(Type, Pc.cod_type_id == Type.cod_type)\
                .join(State, Pc.cod_state_id == State.cod_state)\
                .join(Employes, Pc.cod_pc == Employes.cod_pc_id, isouter=True)\
                .order_by(
                    db.case(
                    (Pc.cod_state_id == 2,0),
                    (Pc.cod_state_id == 3,1),
                    (Pc.cod_state_id == 1,2),
                    (Pc.cod_state_id == 4,3),
                    (Pc.cod_state_id == 3,4),
                    else_=5))\
                .order_by(Pc.date_received.desc())\
                .all()

    data = []

    for computer in resultado:
        data.append({
            'cod_pc': computer.cod_pc,
            'name_computer': computer.name_computer,
            'service_tag': computer.service_tag,
            'name_state': computer.name_state,
            'cod_state' : computer.cod_state,
            'date_received': computer.date_received,
            'cod_model' : computer.cod_model,
            'name_model': computer.name_model,
            'cod_brand' : computer.cod_brand,
            'name_brand': computer.name_brand,
            'cod_type'  : computer.cod_type,
            'name_type': computer.name_type,
            'name_user': computer.lastname_user
            
        })       

    return jsonify(data)

@app.route('/get_history', methods = ['GET'])
def get_historial():

    resultado = db.session.query(
                Log.cod_log,
                Log.log_pc_id,
                Log.log_pc_nc,
                Log.log_pc_st,
                Log.log_date,
                Log.log_state,
                Log.log_archivo,
                Log.log_cod_user_id,
                Log.log_cod_employe,
                Log.log_name_employe
                )\
                .order_by(Log.log_date.desc())\
                .order_by(Log.log_pc_id.desc())\
                .all()

    data = []

    for d in resultado:
        data.append({
            'cod_log':d.cod_log,
            'log_pc_id':d.log_pc_id,
            'log_pc_nc':d.log_pc_nc,
            'log_pc_st':d.log_pc_st,
            'log_date':d.log_date,
            'log_state':d.log_state,
            'log_archivo':d.log_archivo,
            'log_cod_user_id':d.log_cod_user_id,
            'log_cod_employe':d.log_cod_employe,
            'log_name_employe':d.log_name_employe
        })

    return jsonify(data)

@app.route('/listado', methods = ['GET','POST'])
@jwt_required()
def listado():
    current_user_id = get_jwt_identity()
    current_user = Users.query.get(current_user_id)
    return f'Hello, {current_user.nick_name}!'

@app.route('/get_noasignado',  methods = ['GET'])
def get_noasignado():

    resultado = db.session.query(Pc, Pc.service_tag,State.cod_state,Pc.name_computer)\
                .join(State, Pc.cod_state_id == State.cod_state)\
                .filter(State.cod_state == 2)\
                .all()
    data = []
    for d in resultado:
        data.append({ 
                    'state' : d.cod_state,
                    'service_tag' : d.service_tag,
                    'name_computer' : d.name_computer
                    })
    return jsonify(data)

@app.route('/inform_excel',  methods = ['GET'])
def informe():
    output = io.BytesIO()
    wb = Workbook()
    hoja = wb.active

    resultado = db.session.query(Employes.gps_id, Employes.lastname_user, Employes.cod_pc_id, Pc.name_computer, Pc.service_tag, Brand.name_brand, Model.name_model, Type.name_type, Employes.date_delivery, State.name_state)\
        .join(Pc, Pc.cod_pc == Employes.cod_pc_id, isouter=True)\
        .join(Brand, Pc.cod_brand_id == Brand.cod_brand, isouter=True)\
        .join(Model, Pc.cod_model_id == Model.cod_model, isouter=True)\
        .join(Type, Pc.cod_type_id == Type.cod_type, isouter=True)\
        .join(State, Pc.cod_state_id == State.cod_state, isouter=True)\
        .order_by(
            db.case(
                (Pc.cod_state_id == 1, 0),
                (Employes.cod_pc_id == None, 1),
                else_=2))\
        .all()

    headers = ['GPS ID', 'NOMBRE EMPLEADO', 'NOMBRE COMPUTADOR', 'SERVICE TAG', 'MARCA', 'MODELO', 'TIPO', 'FECHA ENTREGA', 'ESTADO']

    hoja.append(headers)

    # Ajustar el ancho de las columnas
    for col_num, value in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        hoja.column_dimensions[column_letter].width = max(len(str(value)) + 5, 15)

    for computer in resultado:
        hoja.append([
            computer.gps_id,
            computer.lastname_user,
            computer.name_computer,
            computer.service_tag,
            computer.name_brand,
            computer.name_model,
            computer.name_type,
            computer.date_delivery,
            computer.name_state
        ])

    wb.save(output)
    output.seek(0)

    return Response(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment;filename=afn.xlsx"})

@app.route('/informe_log',methods = ['GET'])
def informe_log():
    output = io.BytesIO()
    wb = Workbook()
    hoja = wb.active

    resultado = db.session.query(Log.log_pc_id,Log.log_pc_nc,Log.log_pc_st,Log.log_date,Log.log_state,Log.log_cod_employe,Log.log_name_employe,Users.nick_name)\
            .join(Users, Log.log_cod_user_id == Users.cod_user)\
             .order_by(Log.log_date.desc())\
            .order_by(Log.log_pc_id.desc())\
            .all()


    headers = ['CODIGO PC', 'NOMBRE COMPUTADOR', 'SERVICE TAG', 'FECHA', 'ESTADO', 'USUARIO CREADOR', 'GPS ID', 'NOMBRE EMPLEADO']

    hoja.append(headers)

    # Ajustar el ancho de las columnas
    for col_num, value in enumerate(headers, 1):
        column_letter = get_column_letter(col_num)
        hoja.column_dimensions[column_letter].width = max(len(str(value)) + 5, 20)

    for log_afn in resultado:
        hoja.append([
            log_afn.log_pc_id,
            log_afn.log_pc_nc,
            log_afn.log_pc_st,
            log_afn.log_date,
            log_afn.log_state,
            log_afn.nick_name,
            log_afn.log_cod_employe,
            log_afn.log_name_employe
        ])

    wb.save(output)
    output.seek(0)

    return Response(output, mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment;filename=afn_log.xlsx"})

@app.route('/consulta')
def consulta():

    resultado = db.session.query(Log,)\
            .order_by(Log.log_pc_id,Log.log_date.desc())\
            .all()
    
    data = []

    for computer in resultado:
        data.append({
            'log_pc_id': computer.log_pc_id,
            'log_pc_nc': computer.log_pc_nc,
            'log_pc_st': computer.log_pc_st,
            'log_date': computer.log_date,
            'log_state': computer.log_state,
            'log_cod_user_id': computer.log_cod_user_id,
            'log_cod_employe': computer.log_cod_employe,
            'log_name_employe': computer.log_name_employe
        })       

    return jsonify(data)

@app.route('/datos_prueba')
def addu():

    new_type = Type(name_type='notebook')
    new_profile = Profiles(name='administrador')
    new_profile2 = Profiles(name='usuario')
    new_brand = Brand(name_brand='dell')
    new_model = Model(name_model='modelo1')
    new_model2 = Model(name_model='modelo2')
    new_model3 = Model(name_model='modelo3')
    new_state2 = State(name_state='asignado')
    new_state3 = State(name_state='disponible')
    new_state4 = State(name_state='reparación')
    new_state5 = State(name_state='robado')
    new_state6 = State(name_state='dar de baja')
    new_user_profile = Users_profiles(cod_user_id=1,cod_profile_id=1)

    new_user = Users(nick_name='javier', u_password='$2b$12$9q6oTe2dQoV/YdPIfrVAZeFA4P3fZYdHfQzHmZkeatX2dJto/eiRW')
    db.session.add(new_user)
    db.session.add(new_type)
    db.session.add(new_profile)
    db.session.add(new_profile2)
    db.session.add(new_brand)
    db.session.add(new_model)
    db.session.add(new_model2)
    db.session.add(new_model3)
    db.session.add(new_user_profile)
    db.session.add(new_state2)
    db.session.add(new_state3)
    db.session.add(new_state4)
    db.session.add(new_state5)
    db.session.add(new_state6)
    db.session.commit()

    return jsonify({'datos':'Cargados'})

###prueba datos random
@app.route('/prueba_datos')
def prueba_datos():

    names = [fake.unique.first_name() for i in range(500)]
    assert len(set(names)) == len(names)
    numberandom = [fake.unique.random_int() for i in range(500)]
    assert len(set(numberandom)) == len(numberandom)

    #for _ in range(1):
    #    new_dato = Prueba(gps_id=numberandom,lastname_user=names,create_date=fake.date_time_this_year(),cod_employe_id=1)
    #    db.session.add(new_dato)
    #    db.session.commit()


    #for _ in range(30):
    #    new_dato = Pc(name_computer=fake.unique.license_plate(),service_tag=fake.unique.random_int(),date_received=fake.date_time_this_year(),cod_model_id=1,cod_brand_id=1,cod_type_id=1,cod_user_id=1,cod_state_id=fake.random_int(max=5,min=1))
    #    db.session.add(new_dato)
    #    db.session.commit()

    return jsonify({"datos":"random"})

@app.route('/ExcelData', methods=['GET', 'POST'])
def excelData():
    if request.method == 'POST':
        if 'excel_file' not in request.files:
            return jsonify({"mensaje": "No se adjuntó archivo"}), 401

        excel_file = request.files['excel_file']

        if excel_file.filename == '':
            return jsonify({"mensaje": "No se adjuntó archivo"}), 401

        if excel_file:
            wb = load_workbook(excel_file)
            sheet = wb.active

            #agregar pc
            for row in sheet.iter_rows(min_row=2):
                brand_name = str(row[4].value)
                model_name = str(row[5].value)
                type_name = str(row[6].value)

                search_brand = Brand.query.filter_by(name_brand=brand_name).first()
                search_model = Model.query.filter_by(name_model=model_name).first()
                search_type = Type.query.filter_by(name_type=type_name).first()

                if search_brand:
                    brand = search_brand.obtenerCod()
                else:
                    brand = None

                if search_model:
                    modelo = search_model.obtenerCod()
                else:
                    modelo = None

                if search_type:
                    tipo = search_type.obtenerCod()
                else:
                    tipo = None
                

                exist_service_tag = str(row[3].value)
                exist_st = Pc.query.filter_by(service_tag=exist_service_tag).first()
                name_computer = str(row[2].value)
                existe_name_computer = Pc.query.filter_by(name_computer=name_computer).first()
                #print(exist_st,' GPS ID: '+str(gps_id),' Full Name: '+str(lastname_user),' NC: '+str(name_computer),' ST: '+str(exist_service_tag),'Marca :'+str(brand),' Modelo :'+str(modelo)+' Tipo :'+str(tipo))

                if exist_service_tag and name_computer:
                    if exist_st is None and existe_name_computer is None:
                        new_computer = Pc(name_computer=name_computer, service_tag=exist_service_tag,cod_brand_id=brand, cod_state_id=2, cod_type_id=tipo, cod_model_id=modelo)
                        db.session.add(new_computer)
                        db.session.commit()
                    else:
                        #print(exist_service_tag+' duplicado')
                        #print(name_computer+' duplicado')
                        exist_st.name_computer = name_computer
                        exist_st.service_tag = exist_service_tag
                        exist_st.cod_brand_id = brand
                        exist_st.cod_type_id = tipo
                        exist_st.cod_model_id = modelo
                        print('Modificado '+str(brand))
                        db.session.commit()
            
            #agregar empleado
            for row in sheet.iter_rows(min_row=2):
                gps_id = row[0].value
                lastname_user = row[1].value
                exist_gps = Employes.query.filter_by(gps_id=gps_id).first()
                if gps_id and lastname_user:
                    if exist_gps is None:
                        new_employe = Employes(gps_id=gps_id,lastname_user=lastname_user,cod_employe_id=1)
                        db.session.add(new_employe)
                        db.session.commit()
                    else:
                        exist_gps.gps_id = gps_id
                        exist_gps.lastname_user = lastname_user
                        db.session.commit()

            #asignar pc
            for row in sheet.iter_rows(min_row=2):
                asignado = row[8].value
                gps = row[0].value
                st = str(row[3].value)

                existe_employe = Employes.query.filter_by(gps_id=gps).first()
                existe_st = Pc.query.filter_by(service_tag=st).first()

                if existe_employe and asignado:
                    cod = existe_st.cod_pc
                    existe_employe.cod_pc_id = cod
                    existe_st.cod_state_id = asignado
                    db.session.commit()
                    #print('PC asignado...')
                    #print('GPS :'+str(gps)+' asignado :'+str(asignado)+' st :'+str(st)+'/// ST:'+str(st))

            db.session.close()
            return jsonify({"mensaje": "Datos ingresados"})

    return jsonify({"mensaje": "Error conexión"}), 402


with app.app_context():
    db.create_all()

if __name__ == '__main__':
    hostname = socket.gethostname()
    ip_address = socket.gethostbyname(hostname)
    app.run(host=hostname,port=8000,debug=True)
    ##app.run(debug=True)